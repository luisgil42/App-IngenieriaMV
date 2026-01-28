from __future__ import annotations

import io
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from usuarios.decoradores import rol_requerido

from .forms_deals import DealCreateForm, DealForm, DealStageForm
from .models import Company, Deal, DealAttachment, DealStage
from .permissions_ui import can_delete_comercial, ensure_finanzas_access

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None


def _ensure_access(request) -> bool:
    return ensure_finanzas_access(request)


def _get_or_create_default_stage_en_proceso(request) -> DealStage:
    """
    Retorna la etapa por defecto 'En proceso'.
    Si no existe, la crea activa con sort_order=0.
    """
    st = DealStage.objects.filter(name__iexact="En proceso").first()
    if st:
        return st

    st = DealStage.objects.filter(name__icontains="proceso").order_by("id").first()
    if st:
        return st

    st = DealStage.objects.create(
        name="En proceso",
        sort_order=0,
        is_active=True,
        created_by=request.user,
    )
    return st


def _save_deal_attachments(request, deal: Deal, *, field_name: str, category: str):
    """
    Guarda adjuntos para un Deal desde uno o varios <input type="file" name="...">.
    No elimina existentes: solo agrega nuevos.
    """
    if request.method != "POST":
        return

    files = request.FILES.getlist(field_name) or []
    if not files:
        return

    saved_any = False
    for f in files:
        try:
            if not f:
                continue
            DealAttachment.objects.create(
                deal=deal,
                category=category,
                file=f,
                original_name=getattr(f, "name", "") or "",
                content_type=getattr(f, "content_type", "") or "",
                size_bytes=getattr(f, "size", 0) or 0,
                uploaded_by=request.user,
            )
            saved_any = True
        except Exception:
            continue

    if saved_any:
        Deal.objects.filter(id=deal.id).update(last_activity_at=timezone.now())


# ---------------------------------------------------------
# ✅ NEGOCIOS: ELIMINAR ADJUNTO (INICIAL/ANALISIS)
# ---------------------------------------------------------
@login_required
def deal_attachment_delete(request, pk: int):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    if request.method != "POST":
        messages.error(request, "Acción inválida.")
        return redirect("finanzas_comercial:deal_list")

    att = get_object_or_404(DealAttachment, pk=pk)
    deal_id = att.deal_id

    # (opcional) regla de permisos extra: solo creador / superuser / comercial jefe
    # Por ahora: si tiene acceso al módulo, puede borrar adjuntos del deal.
    try:
        # eliminar archivo físico
        try:
            if att.file:
                att.file.delete(save=False)
        except Exception:
            pass

        att.delete()
        Deal.objects.filter(id=deal_id).update(last_activity_at=timezone.now())
        messages.success(request, "✅ Adjunto eliminado.")
    except Exception:
        messages.error(request, "No se pudo eliminar el adjunto.")

    next_url = (request.POST.get("next") or "").strip()
    if next_url:
        return redirect(next_url)
    return redirect("finanzas_comercial:deal_edit", pk=deal_id)


# ---------------------------------------------------------
# NEGOCIOS: LISTA (Excel filters + columnas + paginación)
# ---------------------------------------------------------
@login_required
def deal_list(request):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    show_stages = (request.GET.get("stages") or "").strip() == "1"

    # cantidad / page
    cantidad = (request.GET.get("cantidad") or "20").strip()
    if cantidad not in ("5", "10", "20", "50", "100"):
        cantidad = "20"

    # quick action: cambio rápido de etapa (mantener)
    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "quick_stage":
            deal_id = request.POST.get("deal_id")
            new_stage_id = request.POST.get("stage_id")
            if deal_id and new_stage_id:
                Deal.objects.filter(id=deal_id).update(stage_id=new_stage_id, last_activity_at=timezone.now())
                messages.success(request, "Etapa actualizada.")
            return redirect(request.get_full_path() or "finanzas_comercial:deal_list")

    deals_qs = (
        Deal.objects
        .select_related("company", "owner", "stage")
        .all()
        .order_by("-created_at", "-id")
    )

    # Prefetch de adjuntos
    att_qs = (
        DealAttachment.objects
        .only("id", "deal_id", "category", "file", "original_name", "uploaded_at")
        .order_by("-uploaded_at", "-id")
    )
    deals_qs = deals_qs.prefetch_related(Prefetch("attachments", queryset=att_qs, to_attr="prefetched_attachments"))

    stages = DealStage.objects.filter(is_active=True).order_by("sort_order", "name", "id")
    companies = Company.objects.filter(is_active=True).order_by("name", "id")

    rows = []
    for d in deals_qs:
        atts = getattr(d, "prefetched_attachments", []) or []

        initial_files = []
        analyzed_files = []

        for a in atts:
            cat = (getattr(a, "category", "") or "").upper().strip()
            if cat == DealAttachment.Category.INICIAL:
                initial_files.append(a)
            elif cat == DealAttachment.Category.ANALISIS:
                analyzed_files.append(a)

        d.initial_files = initial_files
        d.initial_count = len(initial_files)

        d.analyzed_files = analyzed_files
        d.analyzed_count = len(analyzed_files)

        rows.append(d)

    paginator = Paginator(rows, int(cantidad))
    page_number = request.GET.get("page") or 1
    pagina = paginator.get_page(page_number)

    # base_qs para links de paginación (sin page)
    qd = request.GET.copy()
    if "page" in qd:
        qd.pop("page")
    base_qs = qd.urlencode()

    can_delete = request.user.is_superuser  # ajusta si aplica

    return render(request, "finanzas_comercial/deal_list.html", {
        "pagina": pagina,
        "cantidad": cantidad,
        "base_qs": base_qs,
        "rows": pagina,  # por compat si tu template usa rows
        "stages": stages,
        "companies": companies,
        "show_stages": show_stages,
        "can_delete": can_delete,
    })


# -----------------
# NEGOCIOS: CREATE (SIN PREGUNTAR ETAPA)
# -----------------
@login_required
def deal_create(request):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    form = DealCreateForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)

        # ✅ Etapa por defecto: En proceso
        obj.stage = _get_or_create_default_stage_en_proceso(request)

        obj.created_by = request.user
        obj.last_activity_at = timezone.now()
        obj.save()

        # ✅ Adjuntos iniciales
        _save_deal_attachments(
            request,
            obj,
            field_name="initial_attachments",
            category=DealAttachment.Category.INICIAL,
        )

        messages.success(request, "✅ Negocio creado (Etapa: En proceso).")
        return redirect("finanzas_comercial:deal_list")

    return render(request, "finanzas_comercial/deal_form.html", {
        "title": "Crear negocio",
        "form": form,
        "is_edit": False,
        "obj": None,
        "initial_existing": [],
    })


# -----------------
# NEGOCIOS: EDIT
# -----------------
@login_required
def deal_edit(request, pk: int):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    obj = get_object_or_404(Deal, pk=pk)
    form = DealForm(request.POST or None, request.FILES or None, instance=obj)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.last_activity_at = timezone.now()
        obj.save()

        # ✅ Si suben más archivos INICIALES al editar
        _save_deal_attachments(
            request,
            obj,
            field_name="initial_attachments",
            category=DealAttachment.Category.INICIAL,
        )

        messages.success(request, "✅ Negocio actualizado.")
        return redirect("finanzas_comercial:deal_list")

    initial_existing = DealAttachment.objects.filter(
        deal=obj,
        category=DealAttachment.Category.INICIAL
    ).order_by("-uploaded_at", "-id")

    return render(request, "finanzas_comercial/deal_form.html", {
        "title": "Editar negocio",
        "form": form,
        "is_edit": True,
        "obj": obj,
        "initial_existing": list(initial_existing),
    })


# -----------------
# NEGOCIOS: DELETE (solo Comercial_Jefe / Superuser)
# -----------------
@login_required
@rol_requerido("Comercial_Jefe")
def deal_delete(request, pk: int):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    if not (can_delete_comercial(request.user) or request.user.is_superuser):
        messages.error(request, "No tienes permisos para eliminar negocios.")
        return redirect("finanzas_comercial:deal_list")

    obj = get_object_or_404(Deal, pk=pk)

    if request.method != "POST":
        messages.error(request, "Acción inválida.")
        return redirect("finanzas_comercial:deal_list")

    try:
        obj.delete()
        messages.success(request, "✅ Negocio eliminado.")
    except Exception:
        messages.error(request, "No se pudo eliminar el negocio.")

    return redirect("finanzas_comercial:deal_list")


# ---------------------------------------------------------
# ✅ NEGOCIOS: SUBIR ARCHIVOS ANALIZADOS (separados)
# ---------------------------------------------------------
@login_required
def deal_analysis_upload(request, pk: int):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    obj = get_object_or_404(Deal, pk=pk)
    next_url = (request.GET.get("next") or "").strip()

    if request.method == "POST":
        _save_deal_attachments(
            request,
            obj,
            field_name="analysis_attachments",
            category=DealAttachment.Category.ANALISIS,
        )
        messages.success(request, "✅ Archivos de análisis cargados.")
        return redirect(next_url or "finanzas_comercial:deal_list")

    analyzed = DealAttachment.objects.filter(
        deal=obj,
        category=DealAttachment.Category.ANALISIS
    ).order_by("-uploaded_at", "-id")

    return render(request, "finanzas_comercial/deal_analysis_upload.html", {
        "title": "Cargar archivos analizados",
        "obj": obj,
        "analyzed": analyzed,
        "next_url": next_url,
    })


# -------------------------
# ETAPAS: CREATE/EDIT/DELETE
# -------------------------
@login_required
@rol_requerido("Comercial_Jefe")
def deal_stage_create(request):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    if request.method != "POST":
        return redirect("finanzas_comercial:deal_list")

    form = DealStageForm(request.POST or None)
    if form.is_valid():
        st = form.save(commit=False)
        st.created_by = request.user
        st.save()
        messages.success(request, "✅ Etapa creada.")
    else:
        messages.error(request, "No se pudo crear la etapa. Revisa el nombre (debe ser único).")

    return redirect(f"{reverse('finanzas_comercial:deal_list')}?stages=1")


@login_required
@rol_requerido("Comercial_Jefe")
def deal_stage_edit(request, pk: int):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    st = get_object_or_404(DealStage, pk=pk)

    if request.method != "POST":
        return redirect(f"{reverse('finanzas_comercial:deal_list')}?stages=1")

    form = DealStageForm(request.POST or None, instance=st)
    if form.is_valid():
        form.save()
        messages.success(request, "✅ Etapa actualizada.")
    else:
        messages.error(request, "No se pudo actualizar la etapa.")

    return redirect(f"{reverse('finanzas_comercial:deal_list')}?stages=1")


@login_required
@rol_requerido("Comercial_Jefe")
def deal_stage_delete(request, pk: int):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    st = get_object_or_404(DealStage, pk=pk)

    if request.method != "POST":
        return redirect(f"{reverse('finanzas_comercial:deal_list')}?stages=1")

    try:
        st.delete()
        messages.success(request, "✅ Etapa eliminada.")
    except Exception:
        messages.error(request, "No se pudo eliminar (puede estar en uso por negocios).")

    return redirect(f"{reverse('finanzas_comercial:deal_list')}?stages=1")


# -------------------------
# EXPORT NEGOCIOS (xlsx)
# -------------------------
@login_required
def deal_export_xlsx(request):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    if Workbook is None:
        messages.error(request, "openpyxl no está disponible en el entorno.")
        return redirect("finanzas_comercial:deal_list")

    q = (request.GET.get("q") or "").strip()
    stage_id = (request.GET.get("stage") or "").strip()
    company_id = (request.GET.get("company") or "").strip()
    active = (request.GET.get("active") or "").strip()

    qs = Deal.objects.select_related("stage", "company", "owner", "created_by").all().order_by("-created_at", "-id")

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(company__name__icontains=q) | Q(owner__email__icontains=q))
    if stage_id.isdigit():
        qs = qs.filter(stage_id=int(stage_id))
    if company_id.isdigit():
        qs = qs.filter(company_id=int(company_id))
    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Negocios"

    headers = [
        "Nombre del negocio",
        "Etapa",
        "Fecha de cierre (YYYY-MM-DD HH:MM)",
        "Empresa (ID)",
        "Empresa",
        "Propietario (email)",
        "Valor",
        "Activo (SI/NO)",
        "Creado por (email)",
        "Fecha creación (YYYY-MM-DD HH:MM)",
    ]
    ws.append(headers)

    for d in qs.iterator():
        ws.append([
            d.name or "",
            (d.stage.name if d.stage else ""),
            (d.close_at.strftime("%Y-%m-%d %H:%M") if d.close_at else ""),
            (d.company.id if d.company else ""),
            (d.company.name if d.company else ""),
            (d.owner.email if d.owner else ""),
            str(d.value or Decimal("0")),
            "SI" if d.is_active else "NO",
            (d.created_by.email if d.created_by else ""),
            (d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else ""),
        ])

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"negocios_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# -------------------------
# IMPORT NEGOCIOS (xlsx)
# -------------------------
@login_required
def deal_import(request):
    if not _ensure_access(request):
        return redirect("core:dashboard")

    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Debes seleccionar un archivo .xlsx")
            return redirect("finanzas_comercial:deal_import")

        if openpyxl is None:
            messages.error(request, "openpyxl no está disponible en el entorno.")
            return redirect("finanzas_comercial:deal_list")

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Archivo inválido. Debe ser .xlsx")
            return redirect("finanzas_comercial:deal_import")

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            messages.error(request, "El archivo no tiene datos.")
            return redirect("finanzas_comercial:deal_import")

        header = [str(x or "").strip().lower() for x in rows[0]]

        def col(name: str):
            name = name.strip().lower()
            return header.index(name) if name in header else None

        idx_name = col("nombre del negocio")
        idx_stage = col("etapa")
        idx_close = col("fecha de cierre (yyyy-mm-dd hh:mm)") or col("fecha de cierre")
        idx_company_id = col("empresa (id)")
        idx_company_name = col("empresa")
        idx_owner_email = col("propietario (email)")
        idx_value = col("valor")
        idx_active = col("activo (si/no)")

        if idx_name is None or idx_stage is None:
            messages.error(request, "Formato incorrecto: faltan columnas requeridas (Nombre del negocio, Etapa).")
            return redirect("finanzas_comercial:deal_import")

        created = 0
        updated = 0
        skipped = 0

        stage_cache = {}
        company_cache = {}

        for r in rows[1:]:
            name = (str(r[idx_name] or "").strip() if idx_name is not None else "")
            stage_name = (str(r[idx_stage] or "").strip() if idx_stage is not None else "")

            if not name:
                continue
            if not stage_name:
                skipped += 1
                continue

            s_key = stage_name.strip().lower()
            stage = stage_cache.get(s_key)
            if stage is None:
                stage = DealStage.objects.filter(name__iexact=stage_name).first()
                if stage is None:
                    stage = DealStage.objects.create(
                        name=stage_name,
                        is_active=True,
                        sort_order=0,
                        created_by=request.user
                    )
                stage_cache[s_key] = stage

            company = None
            c_id_raw = (str(r[idx_company_id] or "").strip() if idx_company_id is not None else "")
            c_name = (str(r[idx_company_name] or "").strip() if idx_company_name is not None else "")

            if c_id_raw.isdigit():
                company = Company.objects.filter(id=int(c_id_raw)).first()
            if company is None and c_name:
                key = c_name.strip().lower()
                company = company_cache.get(key)
                if company is None:
                    company = Company.objects.filter(name__iexact=c_name).first()
                    if company is None:
                        company = Company.objects.create(name=c_name, is_active=True, created_by=request.user)
                    company_cache[key] = company

            owner = None
            owner_email = (str(r[idx_owner_email] or "").strip().lower() if idx_owner_email is not None else "")
            if owner_email:
                from .forms import comercial_users_qs
                owner = comercial_users_qs().filter(email__iexact=owner_email).first()

            raw_val = r[idx_value] if idx_value is not None else 0
            try:
                value = Decimal(str(raw_val).replace(",", "."))
            except Exception:
                value = Decimal("0")

            activo_raw = (str(r[idx_active] or "").strip().upper() if idx_active is not None else "SI")
            is_active = (activo_raw != "NO")

            close_at = None
            if idx_close is not None:
                c = r[idx_close]
                try:
                    if c:
                        close_at = timezone.make_aware(c) if timezone.is_naive(c) else c
                except Exception:
                    close_at = None

            try:
                obj = Deal.objects.filter(name__iexact=name, company=company).first()

                if obj:
                    obj.stage = stage
                    obj.close_at = close_at
                    obj.owner = owner
                    obj.value = value
                    obj.is_active = is_active
                    obj.last_activity_at = timezone.now()
                    obj.save()
                    updated += 1
                else:
                    Deal.objects.create(
                        name=name,
                        stage=stage,
                        close_at=close_at,
                        company=company,
                        owner=owner,
                        value=value,
                        is_active=is_active,
                        created_by=request.user,
                        last_activity_at=timezone.now(),
                    )
                    created += 1
            except Exception:
                skipped += 1

        messages.success(request, f"✅ Importación lista. Creados: {created} | Actualizados: {updated} | Omitidos: {skipped}")
        return redirect("finanzas_comercial:deal_list")

    return render(request, "finanzas_comercial/deal_import.html")