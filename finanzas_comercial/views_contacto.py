import io
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from common.permissions import require_perm
# finanzas_comercial/views.py (o donde la tengas)
from usuarios.decoradores import user_has_role  # usa tu helper de roles (M2M)
from usuarios.decoradores import rol_requerido

from .forms import ContactForm, comercial_users_qs
from .models import Company, Contact
from .permissions_ui import can_delete_finanzas, ensure_finanzas_access

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None

@login_required
def index(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    # Home del módulo: manda a contactos (o cambia a companies/deals si quieres)
    return redirect("finanzas_comercial:contact_list")

@login_required
def contact_list(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    q = (request.GET.get("q") or "").strip()
    owner_id = (request.GET.get("owner") or "").strip()
    active = (request.GET.get("active") or "").strip()

    cantidad = (request.GET.get("cantidad") or "20").strip()
    if cantidad not in ("5", "10", "20", "50", "100"):
        cantidad = "20"

    page = (request.GET.get("page") or "1").strip()

    qs = (
        Contact.objects.select_related("company", "owner")
        .all()
        .order_by("-created_at", "-id")
    )

    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
            | Q(phone__icontains=q)
            | Q(job_title__icontains=q)
            | Q(company__name__icontains=q)
        )

    if owner_id.isdigit():
        qs = qs.filter(owner_id=int(owner_id))

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    paginator = Paginator(qs, int(cantidad))
    pagina = paginator.get_page(page)

    # base_qs: todos los params menos page (para links de paginación)
    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    base_qs = params.urlencode()

    can_delete = can_delete_finanzas(request.user)

    return render(
        request,
        "finanzas_comercial/contact_list.html",
        {
            "pagina": pagina,
            "cantidad": cantidad,
            "base_qs": base_qs,
            "q": q,
            "owner_id": owner_id,
            "active": active,
            "can_delete": can_delete,
        },
    )

@login_required
def contact_create(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    # ✅ IMPORTANTE: request.FILES (por tu Contact.logo / Company.logo)
    form = ContactForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✅ Contacto creado.")
        return redirect("finanzas_comercial:contact_list")

    return render(
        request,
        "finanzas_comercial/contact_form.html",
        {
            "title": "Crear contacto",
            "form": form,
            "is_edit": False,
            "page_debug": "CONTACT_CREATE",
        },
    )


@login_required
def contact_edit(request, pk: int):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    obj = get_object_or_404(Contact, pk=pk)

    form = ContactForm(request.POST or None, request.FILES or None, instance=obj)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✅ Contacto actualizado.")
        return redirect("finanzas_comercial:contact_list")

    return render(
        request,
        "finanzas_comercial/contact_form.html",
        {
            "title": "Editar contacto",
            "form": form,
            "is_edit": True,
            "obj": obj,
        },
    )

@login_required
@rol_requerido("Comercial_Jefe", "Admin") 
def contact_delete(request, pk: int):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    # ✅ doble seguridad
    if not can_delete_finanzas(request.user):
        messages.error(request, "No tienes permisos para eliminar contactos.")
        return redirect("finanzas_comercial:contact_list")

    obj = get_object_or_404(Contact, pk=pk)

    if request.method != "POST":
        messages.error(request, "Acción inválida.")
        return redirect("finanzas_comercial:contact_list")

    try:
        obj.delete()
        messages.success(request, "✅ Contacto eliminado.")
    except Exception:
        messages.error(request, "No se pudo eliminar el contacto.")

    return redirect("finanzas_comercial:contact_list")
@login_required
def contact_export_xlsx(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    if Workbook is None:
        messages.error(request, "openpyxl no está disponible en el entorno.")
        return redirect("finanzas_comercial:contact_list")

    # Exporta lo mismo que estás viendo (filtros)
    q = (request.GET.get("q") or "").strip()
    owner_id = (request.GET.get("owner") or "").strip()
    active = (request.GET.get("active") or "").strip()

    qs = Contact.objects.select_related("company", "owner").all().order_by("-created_at", "-id")

    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
            | Q(phone__icontains=q)
            | Q(job_title__icontains=q)
            | Q(company__name__icontains=q)
        )

    if owner_id.isdigit():
        qs = qs.filter(owner_id=int(owner_id))

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"

    headers = [
        "Nombre",
        "Apellido",
        "Correo",
        "Número de teléfono",
        "Empresa (ID)",
        "Empresa",
        "Cargo",
        "Propietario (email)",
        "Última actividad",
        "Fecha de creación",
        "LinkedIn",
        "Activo (SI/NO)",
    ]
    ws.append(headers)

    for c in qs.iterator():
        ws.append([
            c.first_name or "",
            c.last_name or "",
            c.email or "",
            c.phone or "",
            (c.company.id if c.company else ""),
            (c.company.name if c.company else ""),
            c.job_title or "",
            (c.owner.email if c.owner else ""),
            (c.last_activity_at.strftime("%Y-%m-%d %H:%M") if c.last_activity_at else ""),
            (c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""),
            c.linkedin_url or "",
            "SI" if c.is_active else "NO",
        ])

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"contactos_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
def contact_import_template_xlsx(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    if Workbook is None:
        messages.error(request, "openpyxl no está disponible en el entorno.")
        return redirect("finanzas_comercial:contact_list")

    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # =========================
    # Hoja 1: Formato
    # =========================
    ws = wb.active
    ws.title = "Formato"

    ws.append([
        "Nombre",
        "Apellido",
        "Correo",
        "Número de teléfono",
        "Empresa (ID)",          # ✅ recomendado (evita duplicados)
        "Empresa",               # ✅ opcional (sirve si no existe la empresa)
        "Cargo",
        "Propietario (email)",
        "LinkedIn",
        "Activo (SI/NO)",
    ])

    # Ejemplo (ID vacío, nombre informado)
    ws.append([
        "Luis",
        "Gil",
        "luis.gil@empresa.cl",
        "+56 9 1234 5678",
        "",
        "Empresa Demo",
        "Jefe de Proyecto",
        "comercial@mv.cl",
        "https://www.linkedin.com/in/luisgil/",
        "SI",
    ])

    # Anchos (opcional)
    widths = {
        "A": 18, "B": 18, "C": 28, "D": 18, "E": 12,
        "F": 28, "G": 18, "H": 26, "I": 32, "J": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # =========================
    # Hoja 2: Empresas
    # =========================
    ws2 = wb.create_sheet("Empresas")
    ws2.append(["ID", "Nombre"])

    companies = Company.objects.order_by("name").values_list("id", "name")
    for cid, cname in companies:
        ws2.append([cid, cname])

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 40

    # =========================
    # Validación dropdown: Empresa (ID)
    # =========================
    max_row = ws2.max_row
    if max_row >= 2:
        dv = DataValidation(
            type="list",
            formula1=f"=Empresas!$A$2:$A${max_row}",
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add("E2:E5000")  # columna E = Empresa (ID)

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="formato_import_contactos.xlsx"'
    return resp


@login_required
def contact_import(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Debes seleccionar un archivo .xlsx")
            return redirect("finanzas_comercial:contact_import")

        if openpyxl is None:
            messages.error(request, "openpyxl no está disponible en el entorno.")
            return redirect("finanzas_comercial:contact_list")

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Archivo inválido. Debe ser .xlsx")
            return redirect("finanzas_comercial:contact_import")

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            messages.error(request, "El archivo no tiene datos.")
            return redirect("finanzas_comercial:contact_import")

        header = [str(x or "").strip().lower() for x in rows[0]]

        def col(name: str):
            name = name.strip().lower()
            return header.index(name) if name in header else None

        def _norm(s: str) -> str:
            # Normaliza para comparar (minúsculas + colapsa espacios)
            return " ".join((s or "").strip().lower().split())

        idx_nombre = col("nombre")
        idx_apellido = col("apellido")
        idx_correo = col("correo")
        idx_tel = col("número de teléfono") if "número de teléfono" in header else col("numero de teléfono")
        idx_empresa_id = col("empresa (id)")
        idx_empresa = col("empresa")
        idx_cargo = col("cargo")
        idx_owner = col("propietario (email)")
        idx_linkedin = col("linkedin")
        idx_activo = col("activo (si/no)")

        required = [idx_nombre, idx_apellido]
        if any(x is None for x in required):
            messages.error(request, "Formato incorrecto: faltan columnas requeridas (Nombre, Apellido).")
            return redirect("finanzas_comercial:contact_import")

        created = 0
        updated = 0
        skipped = 0

        # cache simple para no consultar Company repetidamente por el mismo nombre
        company_cache = {}  # norm_name -> Company|None

        for r in rows[1:]:
            nombre = (str(r[idx_nombre] or "").strip() if idx_nombre is not None else "")
            apellido = (str(r[idx_apellido] or "").strip() if idx_apellido is not None else "")
            if not nombre and not apellido:
                continue

            correo = (str(r[idx_correo] or "").strip().lower() if idx_correo is not None else "")
            tel = (str(r[idx_tel] or "").strip() if idx_tel is not None else "")
            cargo = (str(r[idx_cargo] or "").strip() if idx_cargo is not None else "")
            owner_email = (str(r[idx_owner] or "").strip().lower() if idx_owner is not None else "")
            linkedin = (str(r[idx_linkedin] or "").strip() if idx_linkedin is not None else "")
            activo_raw = (str(r[idx_activo] or "").strip().upper() if idx_activo is not None else "SI")
            is_active = (activo_raw != "NO")

            # =========================
            # Empresa: por ID o por nombre
            # =========================
            company = None

            empresa_id_raw = ""
            if idx_empresa_id is not None:
                empresa_id_raw = str(r[idx_empresa_id] or "").strip()

            empresa_name = ""
            if idx_empresa is not None:
                empresa_name = str(r[idx_empresa] or "").strip()

            # 1) Si viene ID, usarlo
            if empresa_id_raw.isdigit():
                company = Company.objects.filter(id=int(empresa_id_raw)).first()

            # 2) Si no hay ID válido, usar nombre (con normalización)
            if company is None and empresa_name:
                key = _norm(empresa_name)

                if key in company_cache:
                    company = company_cache[key]
                else:
                    # Primero intenta exacto ignorando mayúsculas
                    company = Company.objects.filter(name__iexact=empresa_name).first()

                    # Si no encontró, intenta resolver espacios/casos con una búsqueda acotada
                    if company is None:
                        # Trae candidatas (acotado) y compara normalizado
                        cand = Company.objects.filter(name__icontains=empresa_name[:10]).only("id", "name")[:80]
                        for cnd in cand:
                            if _norm(cnd.name) == key:
                                company = cnd
                                break

                    # 3) Si aun no existe => crear automáticamente
                    if company is None:
                        company = Company.objects.create(name=empresa_name, is_active=True)

                    company_cache[key] = company

            # =========================
            # Owner (solo comerciales)
            # =========================
            owner = None
            if owner_email:
                owner = comercial_users_qs().filter(email__iexact=owner_email).first()

            # =========================
            # Upsert por email si viene
            # =========================
            obj = None
            if correo:
                obj = Contact.objects.filter(email__iexact=correo).first()

            try:
                if obj:
                    obj.first_name = nombre
                    obj.last_name = apellido
                    obj.email = correo or obj.email
                    obj.phone = tel or obj.phone
                    obj.company = company
                    obj.job_title = cargo or None
                    obj.owner = owner
                    obj.linkedin_url = linkedin or None
                    obj.is_active = is_active
                    obj.save()
                    updated += 1
                else:
                    Contact.objects.create(
                        first_name=nombre,
                        last_name=apellido,
                        email=correo or None,
                        phone=tel or None,
                        company=company,
                        job_title=cargo or None,
                        owner=owner,
                        linkedin_url=linkedin or None,
                        is_active=is_active,
                    )
                    created += 1
            except Exception:
                skipped += 1

        messages.success(
            request,
            f"✅ Importación lista. Creados: {created} | Actualizados: {updated} | Omitidos: {skipped}"
        )
        return redirect("finanzas_comercial:contact_list")

    return render(request, "finanzas_comercial/contact_import.html")