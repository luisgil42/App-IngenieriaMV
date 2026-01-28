# finanzas_comercial/company_views.py
import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from usuarios.decoradores import rol_requerido

from .forms import CompanyForm, comercial_users_qs
from .models import Company
from .permissions_ui import can_delete_finanzas, ensure_finanzas_access

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None



@login_required
def company_list(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    q = (request.GET.get("q") or "").strip()
    owner_id = (request.GET.get("owner") or "").strip()
    active = (request.GET.get("active") or "").strip()

    cantidad = (request.GET.get("cantidad") or "20").strip()
    if cantidad not in ("5", "10", "20", "50", "100"):
        cantidad = "20"

    page = (request.GET.get("page") or "1").strip()

    qs = (
        Company.objects.select_related("created_by")
        .all()
        .order_by("-created_at", "-id")
    )

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(rut__icontains=q)
            | Q(city__icontains=q)
            | Q(country_region__icontains=q)
            | Q(sector__icontains=q)
            | Q(phone__icontains=q)
        )

    if owner_id.isdigit():
        qs = qs.filter(created_by_id=int(owner_id))

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    paginator = Paginator(qs, int(cantidad))
    pagina = paginator.get_page(page)

    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    base_qs = params.urlencode()

    owners = comercial_users_qs()
    can_delete = can_delete_finanzas(request.user)

    return render(
        request,
        "finanzas_comercial/company_list.html",
        {
            "pagina": pagina,
            "cantidad": cantidad,
            "base_qs": base_qs,
            "q": q,
            "owner_id": owner_id,
            "active": active,
            "owners": owners,
            "can_delete": can_delete,
        },
    )


@login_required
def company_create(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    form = CompanyForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.created_by = request.user
        obj.save()
        messages.success(request, "✅ Empresa creada.")
        return redirect("finanzas_comercial:company_list")

    return render(request, "finanzas_comercial/company_form.html", {
        "title": "Crear empresa",
        "form": form,
        "is_edit": False,
    })


@login_required
def company_edit(request, pk: int):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    obj = get_object_or_404(Company, pk=pk)
    form = CompanyForm(request.POST or None, request.FILES or None, instance=obj)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✅ Empresa actualizada.")
        return redirect("finanzas_comercial:company_list")

    return render(request, "finanzas_comercial/company_form.html", {
        "title": "Editar empresa",
        "form": form,
        "is_edit": True,
        "obj": obj,
    })


@login_required
@rol_requerido("Comercial_Jefe", "Admin") 
def company_delete(request, pk: int):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    # doble seguro
    if not can_delete_finanzas(request.user):
        messages.error(request, "No tienes permisos para eliminar empresas.")
        return redirect("finanzas_comercial:company_list")

    obj = get_object_or_404(Company, pk=pk)

    if request.method != "POST":
        messages.error(request, "Acción inválida.")
        return redirect("finanzas_comercial:company_list")

    try:
        obj.delete()
        messages.success(request, "✅ Empresa eliminada.")
    except Exception:
        messages.error(request, "No se pudo eliminar la empresa.")

    return redirect("finanzas_comercial:company_list")



@login_required
def company_export_xlsx(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    if Workbook is None:
        messages.error(request, "openpyxl no está disponible en el entorno.")
        return redirect("finanzas_comercial:company_list")

    q = (request.GET.get("q") or "").strip()
    owner_id = (request.GET.get("owner") or "").strip()
    active = (request.GET.get("active") or "").strip()

    qs = Company.objects.select_related("created_by").all().order_by("-created_at", "-id")

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(rut__icontains=q)
            | Q(city__icontains=q)
            | Q(country_region__icontains=q)
            | Q(sector__icontains=q)
            | Q(phone__icontains=q)
        )

    if owner_id.isdigit():
        qs = qs.filter(created_by_id=int(owner_id))

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    headers = [
        "Nombre de la empresa",
        "RUT",
        "Ciudad",
        "País/Región",
        "Rubro/Sector",
        "Número de contacto",
        "Última actividad",
        "Registrado por (email)",
        "Fecha de creación",
        "Activo (SI/NO)",
    ]
    ws.append(headers)

    for c in qs.iterator():
        ws.append([
            c.name or "",
            c.rut or "",
            c.city or "",
            c.country_region or "",
            c.sector or "",
            c.phone or "",
            (c.last_activity_at.strftime("%Y-%m-%d %H:%M") if c.last_activity_at else ""),
            (c.created_by.email if c.created_by else ""),
            (c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""),
            "SI" if c.is_active else "NO",
        ])

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"empresas_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



@login_required
def company_import_template_xlsx(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    if Workbook is None:
        messages.error(request, "openpyxl no está disponible en el entorno.")
        return redirect("finanzas_comercial:company_list")

    wb = Workbook()
    ws = wb.active
    ws.title = "Formato"

    ws.append([
        "Nombre de la empresa",
        "RUT",
        "Ciudad",
        "País/Región",
        "Rubro/Sector",
        "Número de contacto",
        "Registrado por (email)",
        "Activo (SI/NO)",
    ])

    ws.append([
        "Empresa Demo",
        "76.123.456-7",
        "Santiago",
        "Chile",
        "Telecomunicaciones",
        "+56 9 1234 5678",
        "comercial@mv.cl",
        "SI",
    ])

    widths = {"A": 30, "B": 16, "C": 18, "D": 18, "E": 22, "F": 18, "G": 26, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="formato_import_empresas.xlsx"'
    return resp

@login_required
def company_import(request):
    if not ensure_finanzas_access(request):
        return redirect("core:dashboard")

    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Debes seleccionar un archivo .xlsx")
            return redirect("finanzas_comercial:company_import")

        if openpyxl is None:
            messages.error(request, "openpyxl no está disponible en el entorno.")
            return redirect("finanzas_comercial:company_list")

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Archivo inválido. Debe ser .xlsx")
            return redirect("finanzas_comercial:company_import")

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            messages.error(request, "El archivo no tiene datos.")
            return redirect("finanzas_comercial:company_import")

        header = [str(x or "").strip().lower() for x in rows[0]]

        def col(name: str):
            name = name.strip().lower()
            return header.index(name) if name in header else None

        idx_name = col("nombre de la empresa")
        idx_rut = col("rut")
        idx_city = col("ciudad")
        idx_country = col("país/región") if "país/región" in header else col("pais/región")
        idx_sector = col("rubro/sector") if "rubro/sector" in header else col("rubro o sector")
        idx_phone = col("número de contacto") if "número de contacto" in header else col("numero de contacto")
        idx_owner = col("registrado por (email)")
        idx_active = col("activo (si/no)")

        if idx_name is None:
            messages.error(request, "Formato incorrecto: falta columna 'Nombre de la empresa'.")
            return redirect("finanzas_comercial:company_import")

        created = updated = skipped = 0

        for r in rows[1:]:
            name = (str(r[idx_name] or "").strip() if idx_name is not None else "")
            if not name:
                continue

            rut = (str(r[idx_rut] or "").strip() if idx_rut is not None else "")
            city = (str(r[idx_city] or "").strip() if idx_city is not None else "")
            country = (str(r[idx_country] or "").strip() if idx_country is not None else "")
            sector = (str(r[idx_sector] or "").strip() if idx_sector is not None else "")
            phone = (str(r[idx_phone] or "").strip() if idx_phone is not None else "")
            owner_email = (str(r[idx_owner] or "").strip().lower() if idx_owner is not None else "")
            active_raw = (str(r[idx_active] or "").strip().upper() if idx_active is not None else "SI")
            is_active = (active_raw != "NO")

            created_by = None
            if owner_email:
                created_by = comercial_users_qs().filter(email__iexact=owner_email).first()

            obj = None
            if rut:
                obj = Company.objects.filter(rut__iexact=rut).first()
            if obj is None:
                obj = Company.objects.filter(name__iexact=name).first()

            try:
                if obj:
                    obj.name = name
                    obj.rut = rut
                    obj.city = city
                    obj.country_region = country
                    obj.sector = sector
                    obj.phone = phone
                    if created_by:
                        obj.created_by = created_by
                    obj.is_active = is_active
                    obj.save()
                    updated += 1
                else:
                    Company.objects.create(
                        name=name,
                        rut=rut,
                        city=city,
                        country_region=country,
                        sector=sector,
                        phone=phone,
                        created_by=created_by or request.user,
                        is_active=is_active,
                    )
                    created += 1
            except Exception:
                skipped += 1

        messages.success(request, f"✅ Importación lista. Creadas: {created} | Actualizadas: {updated} | Omitidas: {skipped}")
        return redirect("finanzas_comercial:company_list")

    return render(request, "finanzas_comercial/company_import.html")